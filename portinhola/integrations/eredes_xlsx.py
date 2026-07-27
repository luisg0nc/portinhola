"""Parser for E-Redes Balcão Digital consumption exports (XLSX).

Two known layouts share the same data semantics:
- monthly:  Data | Hora | Consumo registado (kW) | Estado
- range:    Contador | Data | Hora | Consumo registado, Ativa (kW) | Estado

Values are average kW over a 15-minute slot; `Hora` is the slot END in
Europe/Lisbon local time. Stored rows are (UTC slot start, kWh, quality).
"""

from datetime import UTC, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
from pydantic import BaseModel

LISBON = ZoneInfo("Europe/Lisbon")

IntervalRow = tuple[datetime, float, str]


class EredesExport(BaseModel):
    cpe: str | None
    rows: list[IntervalRow]


def _to_float(value: object) -> float:
    if isinstance(value, int | float):
        return float(value)
    return float(str(value).replace(".", "").replace(",", "."))


def parse_eredes_xlsx(path: Path) -> EredesExport:
    workbook = openpyxl.load_workbook(path, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    cpe: str | None = None
    header: list[str] | None = None
    col: dict[str, int] = {}
    rows: list[IntervalRow] = []
    seen_local_ends: set[datetime] = set()
    seen_utc: set[datetime] = set()

    for raw in sheet.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in raw]
        if header is None:
            if cells and cells[0] == "CPE" and len(cells) > 1 and cells[1]:
                cpe = cells[1]
            if "Data" in cells and "Hora" in cells:
                header = cells
                col["data"] = cells.index("Data")
                col["hora"] = cells.index("Hora")
                col["estado"] = cells.index("Estado") if "Estado" in cells else -1
                col["kw"] = next(
                    i for i, name in enumerate(cells) if name.startswith("Consumo")
                )
            continue

        if not cells[col["data"]] or not cells[col["hora"]]:
            continue

        date_str = cells[col["data"]]
        hora_str = cells[col["hora"]]
        year, month, day = (int(part) for part in date_str.split("/"))
        hour, minute = (int(part) for part in hora_str.split(":"))

        naive_end = datetime(year, month, day) + timedelta(hours=hour, minutes=minute)  # noqa: DTZ001
        # DST fall-back: the same wall-clock hour appears twice in the file;
        # the second occurrence is the post-transition (fold=1) one.
        fold = 1 if naive_end in seen_local_ends else 0
        seen_local_ends.add(naive_end)
        local_start = (naive_end - timedelta(minutes=15)).replace(tzinfo=LISBON, fold=fold)
        ts_utc = local_start.astimezone(UTC)
        if ts_utc in seen_utc:
            # Residual collision (spring-forward artifact rows) — skip.
            continue
        seen_utc.add(ts_utc)

        kwh = _to_float(cells[col["kw"]]) / 4
        estado = cells[col["estado"]] if col["estado"] >= 0 else "Real"
        quality = "real" if estado == "Real" else "estimated"
        rows.append((ts_utc, kwh, quality))

    workbook.close()
    return EredesExport(cpe=cpe, rows=rows)
